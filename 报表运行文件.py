# -*- coding: utf-8 -*-
"""
Created on Fri Nov  4 11:28:21 2016

@author: JJJ
"""
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart  
from email.mime.application import MIMEApplication 
from WindPy import w 
from dailyReport import dailyReport
import time
import datetime
import openpyxl
import copy
def choices(x):
    while True:
        print ("可选项列表: ")
        for i in range(len(x)):
            print ('\n' + str(i + 1) + '. ' + x[i])
        a = input('\n请按序号输入您的选择(如需输入多个序号,请用逗号分隔): ')
        a = a.split(',')
        try:
            a = [int(i) for i in a if int(i) <= len(x)]
            break
        except:
            pass
    return a
def sendEmail(today, path, title):
    _user = 'shixisheng1@sinvofund.com'
    _pwd = 'Jjj123'
    _to = ['wuliang@sinvofund.com', 'dingping@sinvofund.com', 'yuli@sinvofund.com', 'shixisheng1@sinvofund.com','shenxia@sinvofund.com']
#    _to = ['shixisheng1@sinvofund.com']

    msg = MIMEMultipart(); 
    msg["Subject"] = title + today;
    msg["From"]    = _user;
    msg["To"]      = ','.join(_to)
    if title == '基金':
        titleE = 'Funds'
    elif title == '晨报':
        titleE = 'Morning Report'
    part = MIMEText("尊敬的各位：\n  附件是今天的{},请查收~\n  祝好~\n\n---------------------------------\n新沃基金固定收益部".format(title));  
    msg.attach(part)
    part = MIMEApplication(open(path,'rb').read())
    msg.attach(part)
    part.add_header('Content-Disposition', 'attachment', filename= titleE + today + ".xlsx")     
    try:
        s = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
        s.login(_user, _pwd)
        s.sendmail(_user, _to, msg.as_string())
        s.quit();
        print ("Success!")
    except:
        print ("Falied")
        
if __name__ == '__main__':
    
    start = time.clock();
    w.start();
    today = datetime.date.today()
    today = today.strftime('%Y-%m-%d')
    inputPaths = ['晨报','基金']
#    inputPaths = ['基金']
#    a = choices(inputPaths)
    a = [1,2]
    w.start()
    for j in a:
        i = inputPaths[j-1]
        inputPath = '输入表/' + i + '输入表.xlsx'
        print (i + ' Running....\n')
        if i == '基金':
            names = ['打新基金','长期标准债基(A类)','普通债基(一级A类)','普通债基(二级A类)','灵活配置型(A类)','股票基金']
            book1 = openpyxl.Workbook()
            for sheet in names:
                outputPath = i + '输出/' + sheet + ' ' + today + '.xlsx'
                report = dailyReport(inputPath, sheet, 0, outputPath)
                report.toExcel()
                sheet1 = book1.create_sheet(sheet)
                book2 = openpyxl.load_workbook(outputPath)
                sheet2 = book2.get_active_sheet()
                for row in range(1, sheet2.max_row + 1):
                    for col in range(ord('C'), sheet2.max_column + ord('A')):
                        num1 = chr(col - 2) + str(row)
                        num2 = chr(col) + str(row)
                        sheet1[num1].value = sheet2[num2].value
                        sheet1[num1].font = copy.copy(sheet2[num2].font)
                        sheet1[num1].fill = copy.copy(sheet2[num2].fill)
                        sheet1[num1].border = copy.copy(sheet2[num2].border)
                        sheet1[num1].number_format = copy.copy(sheet2[num2].number_format)
            outputPath = i + '输出/' + i + ' ' + today + '.xlsx'
            book1.remove_sheet(book1.get_sheet_by_name('Sheet'))
            book1.save(outputPath)
            sendEmail(today, outputPath, i)
        else:           
            outputPath = i + '输出/' + i + ' ' + today + '.xlsx'
            report = dailyReport(inputPath, 'Sheet1', 0, outputPath)
            report.toExcel()
            sendEmail(today, outputPath, i)
        end = time.clock()
        print (i + ' Finished....\n')
        print ('run time %fs' % (end - start))
        
    
    w.close();