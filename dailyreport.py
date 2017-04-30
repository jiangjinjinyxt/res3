# -*- coding: utf-8 -*-
"""
Created on Wed Oct 26 09:18:47 2016

@author: JJJ

color codes:
for more imformation, browse https://en.wikipedia.org/wiki/Web_colors#X11_color_names
BLACK = 'FF000000'
WHITE = 'FFFFFFFF'
RED = 'FFFF0000'
DARKRED = 'FF800000'
BLUE = 'FF0000FF'
DARKBLUE = 'FF000080'
GREEN = 'FF00FF00'
DARKGREEN = 'FF008000'
YELLOW = 'FFFFFF00'
DARKYELLOW = 'FF808000'
"""
from WindPy import w
import pandas as pd
import copy
import datetime
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color, numbers

w.start()

add_average_list = ['基金']
add_average_list = [i + '输入表.xlsx' for i in add_average_list]
#function_name
function_names = {'S':'值','V':'值','R':'涨跌幅','RC':'涨跌','MIN':'最小值','MAX':'最大值',
                  'Q1':'1/4分位数','Q2':'1/2分位数','Q3':'3/4分位数',
                  'MAXDRAW':'最大回撤','SHARPE':'年化夏普比率','MEAN':'均值','RANK':'排名',
                  'SPREAD':'基差','PREMIUM':'升贴水 现货对期货'
                  }          
#parameter dictionary 
parameter_dictionary = {'yield':'到期收益率','close':'收盘价','interest rate':'资金价格',
                'nav_adj':'复权净值','mmf_annualizedyield':'七日年化收益','settle':'结算价'
                }
#futures:spot
futures_dictionary = {
    'IF':'000300.SH',
    'IH':'000016.SH',
    'IC':'000905.SH',
    'CU':'S0182161',
    'AL':'S0182162',
    'ZN':'S0048087',
    'RB':'S0033227',
    'PB':'S0048086'
    }
#list of dates in the input excel file
def get_date_list(inputdata):    
    date_list = set()
    columns = [i for i in inputdata.columns if i[:3] == '提取值']
    for i in columns:
        temp = inputdata[i]
        for j in temp:
            try:
                index1 = j.find('(')
                index2 = j.find(',')
                index3 = j.find(')')
                if index2 > 0:
                    start = j[index1 + 1:index2]
                    end = j[index2 + 1:index3]
                    date_list.add(start)
                    date_list.add(end)
                else:
                    start = j[index1 + 1:index3]
                    date_list.add(start)
            except:
                pass
            
    date_list = [i for i in date_list]
    return date_list
    
def get_date_dict(date_list):
    date_dict = {'TODAY':'昨日','YEAR':'今年'}
    for i in date_list:
        try:
            a = i.split('-')[0]
            b = i.split('-')[1]
            if a == 'TODAY':
                date_dict[i] = str(b) + '个交易日'
        except:
            pass
    return date_dict
    
def get_date_before(number_of_days, flag):
    """
    return the date number_of_days before the flag date
    """
    return w.tdaysoffset(-number_of_days, flag).Data[0][0];
      
def get_dates(date_list):
    """
    return dates dictionry 
    """
    TODAY = datetime.date.today() - datetime.timedelta(1)
    TODAY = w.tdaysoffset(0,TODAY).Data[0][0]
    YEAR = datetime.datetime.strptime(str(TODAY.year) + '-01-01', "%Y-%m-%d")
    YEAR = get_date_before(1, YEAR)
    YEAR = w.tdays(beginTime = YEAR, endTime = TODAY).Data[0][0]
    dates = {}
    for date_index in date_list:
        number_of_days = 0
        if str.isdigit(date_index[-1]):
            for i in range(len(date_index)):
                if str.isdigit(date_index[i]):
                    break
            number_of_days = int(date_index[i:])
        dates[date_index] = w.tdaysoffset(-number_of_days, TODAY).Data[0][0]
    dates['YEAR'] = YEAR
    return dates
#function dictionary
def get_func_dictionary(function_names, date_dict, inputdata):
    columns = [i for i in inputdata.columns if i[:3] == '提取值']
    func_dictionary = {'':' '
      }
    for i in columns:
        temp = inputdata[i]
        for j in temp:
            try:
                index1 = j.find('(')
                index2 = j.find(')')
                func = j[:index1]
                a = j[index1 + 1:index2]
                try:
                    [para1, para2] = a.split(',')
                    func_dictionary[j] = date_dict[para1] + '内' + function_names[func]
                except:
                    if a == 'TODAY':
                        func_dictionary[j] = date_dict[a] + function_names[func]
                    else:
                        func_dictionary[j] = date_dict[a] + '前' + function_names[func]
            except:
                pass
    return func_dictionary
        
def get_series(code, parameter, date_index, dates_dictionary):
    """
    codes = code;(code is windcode)    
    fields = parameter;
    dates_index is a string of format like 'begin,end';    
    begin is the begin date of the interval, end is the end date of the interval;
    return the value series in range(begin, end);
    for w.wsd:
        return w.wsd(codes = code, fields = parameter, beginTime = begin, endTime = end).Data[0];
    for w.edb:
        return w.edb(codes = code, beginTime = begin, endTime = end).Data[0];
    """
    [begin,end] = date_index.split(',')
    begin = dates_dictionary[begin]
    end = dates_dictionary[end]
    codes = code.split('-')

    if len(codes) == 1:
        code = codes[0]
        if code.find('.') > 0:
            a = (w.wsd(codes = code, fields = parameter, beginTime = begin, endTime = end)).Data[0]
        else:
            a = (w.edb(codes = code, beginTime = begin, endTime = end)).Data[0]
    elif len(codes) == 2:
        [code1, code2] = codes
        if code1.find('.') > 0:
            a = w.wsd(codes = code1, fields = parameter, beginTime = begin, endTime = end)
        else:
            a = w.edb(codes = code1, beginTime = begin, endTime = end)
        data1 = a.Data[0]
        dates1 = a.Times
        data1 = pd.DataFrame(data1, index = dates1, columns = ['a'])
        if code2.find('.') > 0:
            b = w.wsd(codes = code2, fields = parameter, beginTime = begin, endTime = end)
        else:
            b = w.edb(codes = code2, beginTime = begin, endTime = end)
        data2 = b.Data[0]
        dates2 = b.Times
        data2 = pd.DataFrame(data2, index = dates2, columns = ['b'])
        data = data1.join(data2, how = 'outer')
        data.fillna(method = 'ffill')
        data.fillna(method = 'bfill')
        a = data['a'] - data['b']
    a = pd.Series(a).dropna()
    return a
    
def V(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    function V, returns a value;
    windcode = code;//codes = code
    fields = parameter;
    beginTime = date_index;
    endTime = date_index;
    """
    date_index = dates_dictionary[date_index]
    codes = code.split('-')
    if len(codes) == 1:
        code = codes[0]
        if code.find('.') > 0:
            while True:
                a = (w.wsd(codes = code, fields = parameter, beginTime = date_index, endTime = date_index)).Data[0][0]
                try:
                    int(a)
                    break
                except:
                    date_index = get_date_before(1, date_index)
        else:
            a = (w.edb(codes = code, beginTime = date_index, endTime = date_index)).Data[0][0]
    else:
        [code1,code2] = codes
        if code1.find('.') > 0:
            while True:
                a = (w.wsd(codes = code1, fields = parameter, beginTime = date_index, endTime = date_index)).Data[0][0]
                try:
                    int(a)
                    break
                except:
                    date_index = get_date_before(1, date_index)
        else:
            a = (w.edb(codes = code1, beginTime = date_index, endTime = date_index)).Data[0][0]
        if code2.find('.') > 0:
            while True:
                b = (w.wsd(codes = code2, fields = parameter, beginTime = date_index, endTime = date_index)).Data[0][0]
                try:
                    int(b)
                    break
                except:
                    date_index = get_date_before(1, date_index)
        else:
            b = (w.edb(codes = code2, beginTime = date_index, endTime = date_index)).Data[0][0]
        a = a - b
    return a
    

"""
parameter_type = 1, 是类似于收益率、利率的序列
parameter_type = 2, 是类似于价格的序列
"""    
def Q1(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns 1/4 quantile of the value series
    """
    return get_series(code, parameter, date_index, dates_dictionary).quantile(0.25)
def Q2(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns 1/2 quantile of the value series
    """
    return get_series(code, parameter, date_index, dates_dictionary).quantile(0.5)
def Q3(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns 3/4 quantile of the value series
    """
    return get_series(code, parameter, date_index, dates_dictionary).quantile(0.75)
def MIN(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns minimun of the value series
    """
    return get_series(code, parameter, date_index, dates_dictionary).min()
def MAX(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns maximun of the value series
    """
    return get_series(code, parameter, date_index, dates_dictionary).max()
def MEAN(code, parameter, date_index, parameter_type, dates_dictionary):
    return get_series(code, parameter, date_index, dates_dictionary).mean()
def R(code, parameter, date_index, parameter_type, dates_dictionary):
    c = get_series(code, parameter, date_index, dates_dictionary)
    a = c[c.index[0]]
    b = c[c.index[-1]]
    if parameter_type == 2:
        return (b - a) / a
    else:
        return (b - a) / 100
def RC(code, parameter, date_index, parameter_type, dates_dictionary):
    c = get_series(code, parameter, date_index, dates_dictionary)
    try:
        return c[c.index[-1]] - c[c.index[0]]
    except:
        return 0
def MAXDRAW(code, parameter, date_index, parameter_type, dates_dictionary):
    """
    returns maxdraw in range(date_index[0], date_index[1])
    if parameter_type == 1:
        收益率序列,最大回撤设定为序列的最大绝对值变化
    else:
        价格序列
    """
    date_index = str(date_index)
    values = get_series(code, parameter, date_index, dates_dictionary)
    values.index = range(len(values))
    maxdraw = 0
    if parameter_type == 1:
        for i in range(1,len(values)):
            temp1 = values[:i].max()
            temp2 = values[i]
            if temp2 - temp1 < maxdraw:
                maxdraw = temp2 - temp1
        maxdraw /= 100
    elif parameter_type == 2:
        for i in range(1,len(values)):
            temp1 = values[:i].max()
            temp2 = values[i]
            if (temp2 - temp1) / temp1 < maxdraw:
                maxdraw = (temp2 - temp1) / temp1
    else:
        pass
    return -maxdraw
def SHARPE(code, parameter, date_index, parameter_type, dates_dictionary):
    a = get_series(code, parameter, date_index, dates_dictionary)
    b = ((a - a.shift(1)) / a.shift(1)).dropna()
    try:
        return b.mean() * (252 ** 0.5) / ((b.var() * len(b) / (len(b) - 1)) ** 0.5)
    except ZeroDivisionError:
        print (code + ': Sharpe zero divison')
        return 0

def RANK(code, parameter, date_index, parameter_type, dates_dictionary):
    [begin,end] = date_index.split(',')
    begin = dates_dictionary[begin]
    begin = w.tdaysoffset(1, begin).Data[0][0]
    end = dates_dictionary[end]
    return w.wsd(code, "peer_fund_ return_rank_per", begin, end, "fundType=3;Period=Y;PriceAdj=F").Data[0][0]

def SPREAD(code, parameter, date_index, parameter_type, dates_dictionary):
    futures_price = V(code, parameter, date_index, parameter_type, dates_dictionary)
    code = code.split('.')[0]
    if code == 'T' or code == 'TF':
        date_index = dates_dictionary[date_index]
        date_index = date_index.strftime('%Y-%m-%d')
        main_contract = w.wsd(codes = code + '.CFE', fields = 'trade_hiscode', beginTime = date_index, endTime = date_index).Data[0][0]
        ctd = w.wset('ctd', 'startdate=' + date_index + ';enddate=' + date_index + ';windcode=' + main_contract + ';field=date,ctd_ib').Data[1][0]
        cf = pd.DataFrame(w.wset('conversionfactor','windcode=' + main_contract).Data).T
        cf.index = cf[0]
        try:
            return w.wsd(ctd, 'net_cnbd', date_index, date_index,"credibility=1;PriceAdj=F").Data[0][0] - cf[1][ctd] * futures_price
        except:
            return w.wsd(ctd, 'close', date_index, date_index, "").Data[0][0] - cf[1][ctd] * futures_price
    elif code == 'IF' or code == 'IC' or code == 'IH':
        try:
            spot_price = V(futures_dictionary[code], 'Close', date_index, parameter_type, dates_dictionary)
            return spot_price - futures_price
        except:
            return 0
    else:
        try:
            spot_price = V(futures_dictionary[code], parameter, date_index, parameter_type, dates_dictionary)
            return spot_price - futures_price
        except:
            return 0
def PREMIUM(code, parameter, date_index, parameter_type, dates_dictionary):
    try:
        futures_price = V(code, parameter, date_index, parameter_type, dates_dictionary)
        deliver_date = w.wss(code,'lastdelivery_date').Data[0][0]
        days = (deliver_date - dates_dictionary[date_index]).days
        code = code.split('.')[0]
        if code == 'IF' or code == 'IC' or code == 'IH':
            spot_price = V(futures_dictionary[code], 'Close', date_index, parameter_type, dates_dictionary)
        return (spot_price - futures_price) * 365 / (days * futures_price)
    except:
        return 0
def f(string):
    print (string)
    try:
        index1 = string.find('(')
        if index1 > 0:
            index2 = string.find(')')
            func_name = string[:index1]
            if func_name == 'S':
                func_name = 'V'
            date_index = string[index1+1:index2]
            return [func_name, date_index]
        else:
            return [False, False]
    except:
        return [False, False]
    
def change_nan_to_none(x):
    if type(x) == str:
        return x
    else:
        return ''
def leave80(x):
    x = x.sort_values()
    x.index = range(len(x))
    length = len(x)
    return x[int(length*0.1):int(length*0.9)].mean()
def process_r_data(data, path):
    path = path.split('/')[1]
    columns0 = [column for column in data.columns if column[-1] != ' ']
    columns1 = [column for column in data.columns if column[:3] == '提取值' and column[-1] != ' ']
    columns2 = [column + ' ' for column in columns1]
    [big,small] = ['所属板块-大','所属板块-小']
    data.sort([big,small], inplace = True)
    if path[:2] == '晨报':
        rets_columns = ['债券收益率','期限利差','信用利差']
        for i in rets_columns:
            a = data[data[small] == i]
            data = data[data[small] != i]
            data = data.append(a)
        rets_columns = ['股指期货','国债期货','基本金属期货','能源化工期货','农产品期货']
        for i in rets_columns:
            a = data[data[small] == i]
            data = data[data[small] != i]
            data = data.append(a)
    index_1 = data['所属板块-小'].unique()
    output = pd.DataFrame(columns = data.columns)
    flag_average = False
    if path in add_average_list:
        flag_average = True
    for i in index_1:
        temp1 = copy.deepcopy(data[data['所属板块-小'] == i])
        if flag_average:
            temp2 = pd.DataFrame(columns = columns1, index = [0])
            for co in columns1:
                temp2[co][0] = leave80(temp1[co])
            temp1 = temp1.append(temp2, ignore_index = True)
            temp1.ix[temp1.index[-1],'指标名称'] = '平均值'
            temp1.fillna(method = 'ffill', inplace = True)
        a = pd.DataFrame(copy.deepcopy(temp1.iloc[0,:])).T
        a[columns1] = a[columns2]
        for i in a.index:
            for j in columns1:
                try:
                    a.ix[i,j] = (a.ix[i,j].split(')'))[1]
                except:
                    pass
        a['指标名称'] = a['提取参数']
        b = a.append(temp1, ignore_index = True)
        
        output = output.append(b, ignore_index = True)
    output.drop('提取参数', axis = 1, inplace = 1)
    output.index = range(1, len(output) + 1)
    output1 = copy.deepcopy(output)
    columns0.remove('提取参数')
    output = output[columns0]
    return [output, output1]

def process_raw_data(path, sheetname, index_col):
    data = pd.read_excel(path, sheetname = sheetname, index_col = index_col)
    date_list = get_date_list(data)
    date_dict = get_date_dict(date_list)
    dates_dictionary = get_dates(date_list)
    func_dictionary = get_func_dictionary(function_names, date_dict, data)
    data.index = range(len(data))
    columns0 = ['所属板块-大','所属板块-小','指标名称','提取参数']
    columns1 = [column for column in data.columns if column[:3] == '提取值']
    columns0.extend(columns1)
    columns2 = [column + ' ' for column in columns1]
    columns0.extend(columns2)
    for column in columns1:
        data[column] = data[column].apply(change_nan_to_none)
    for column in columns1:
        data[column + ' '] = data[column]
    for ind in data.index:
        for column in columns1:
            [func_name, date_index] = f(data[column][ind])
            if func_name:             
                code = data.ix[ind,'WIND代码']
                parameter = data.ix[ind,'提取参数']
                parameter_type = data.ix[ind,'提取参数类型']
                print (code)
                data[column][ind] = eval(func_name)(code, parameter, date_index, parameter_type, dates_dictionary)
            else:
                data[column][ind] = np.nan
    def g(x):
        return func_dictionary[x]
    def k(x):
        return parameter_dictionary[x]
    for column in columns2:
        data[column] = data[column] + data[column].apply(g)
    data['提取参数'] = data['提取参数'].apply(k)
    data = data[columns0]
    [data, data1] = process_r_data(data, path)
    return [data, data1]

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    set style for merged cells in cell_range
    """
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment
    rows = ws[cell_range]
    if font:
        first_cell.font = font
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
                
def set_style(cell, cell_type):
    """
    set style for a single cell
    """
    thin = Side(border_style="thin", color="000000")
    font = Font(b=False, color="000000")
    al = Alignment(horizontal="center", vertical="center")
    border = Border(left=thin, right=thin)
    number_format = '0.00'
    if cell_type == 1:
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        fill = PatternFill("solid", fgColor="FFFF00")
    elif cell_type == 2:
        fill = PatternFill("solid", fgColor="FFA07A")
    else:
        fill = PatternFill("solid", fgColor="F5F5F5")
        al = Alignment(horizontal="right", vertical="center")
        if cell_type == 4:
            number_format = '0.00%'
        elif cell_type == 0:
            pass;
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    cell.alignment = al
    if font:
        cell.font = font
    if Border:
        cell.border = cell.border + top
        cell.border = cell.border + bottom
        cell.border = cell.border + left
        cell.border = cell.border + right         
    if fill:
        cell.fill = fill
    if cell_type > 2:
        cell.number_format = number_format

def set_style_for_mergecells(ws, data, index_type):
    if index_type == '所属板块-大':
        index_1 = data[index_type].unique()
        for i in index_1:
            temp = data[data[index_type] == i].index
            begin = temp[0]
            end = temp[-1]
            cell_range = 'A%d:A%d' % (begin,end)
            thin = Side(border_style="thin", color="000000") #黑色单线
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            fill = PatternFill("solid", fgColor= '00BFFF')#蓝底
            font = Font(b=False, color="000000")
            al = Alignment(horizontal="center", vertical="center")
            style_range(ws,cell_range, border = border, fill = fill,font = font, alignment = al)
    elif index_type == '所属板块-小':
        index_2 = data[index_type].unique()
        for j in index_2:
            temp = data[data[index_type] == j].index
            begin = temp[0]
            end = temp[-1]
            cell_range = 'B%d:B%d' % (begin,end)
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            fill = PatternFill("solid", fgColor="3CB371")#绿底
            font = Font(b=False, color="000000")
            al = Alignment(horizontal="center", vertical="center")
            style_range(ws,cell_range, border = border, fill = fill,font = font, alignment = al)
            
def set_style_for_pct(ws, data, data1):
    percent_cell = ['R','MAXDRAW','PREMIUM']
    for i in data.index:
        if data.ix[i,2] in parameter_dictionary.values():
            cell_type = 1;
            for j in range(2,len(data.columns)):
                cell = chr(ord('A') + j) + str(i)
                set_style(ws[cell], cell_type)
        else:
            for j in range(2, len(data.columns)):
                column = data.columns[j] + ' '
                cell = chr(ord('A') + j) + str(i)
                if j == 2:
                    cell_type = 2;
                    set_style(ws[cell], cell_type);
                else:
                    try:
                        a = data1.ix[i, column].split('(')[0]
                        if a in percent_cell:
                            cell_type = 4
                            set_style(ws[cell], cell_type)
                        else:
                            cell_type = 3
                            set_style(ws[cell], cell_type)
                    except:
                        cell_type = 0
                        set_style(ws[cell], cell_type)
    thin = Side(border_style="thin", color="000000");
    i = data.index[-1];
    for j in range(2, len(data.columns)):
        cell = chr(ord('A') + j) + str(i);
        border = Border(bottom=thin);
        bottom = Border(bottom=border.bottom)
        ws[cell].border = ws[cell].border + bottom
def picture(data, data1, outPath):
    [big, small] = ['所属板块-大','所属板块-小'];
    wb = openpyxl.Workbook();
    ws = wb.active;
    for r in dataframe_to_rows(data, index = False, header = False):
        ws.append(r);
    set_style_for_mergecells(ws, data, big)
    set_style_for_mergecells(ws, data, small)
    set_style_for_pct(ws, data, data1)
    wb.save(outPath)
    
class dailyReport(object):
    def __init__(self, inputPath, sheetname, index_col, outputPath):
        [data,data1] = process_raw_data(inputPath, sheetname, index_col)
        self.data1 = data1
        self.data = data
        self.columns = data.columns;
        self.outputPath = outputPath
    def toExcel(self):
        picture(self.data, self.data1, self.outputPath)